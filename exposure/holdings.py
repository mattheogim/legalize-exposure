"""Daily ETF Holdings Snapshot Collector.

Design principle §9: ETFs are connected via holdings, not sector labels.

Three data sources (in priority order):
  1. iShares CSV — daily holdings CSVs from BlackRock (free, ~1-day lag)
  2. SEC N-PORT — quarterly fund holdings filings (free, ~60-day lag)
  3. ETF provider websites — scrape/parse holdings pages (fallback)

The collector:
  - Fetches current holdings for each ETF
  - Maps each holding's company to NAICS codes (via SIC→NAICS crosswalk)
  - Stores daily snapshots in JSON format
  - Maintains a historical archive for point-in-time replay

Directory structure:
    holdings_data/
        snapshots/
            2026-04-11/
                XLE.json
                XLF.json
                ...
            2026-04-10/
                ...
        naics_cache/
            AAPL.json     # ticker → NAICS mapping cache
        metadata.json     # last fetch dates, errors, etc.
"""

from __future__ import annotations

import csv
import io
import json
import logging
import os
import re
import time
import urllib.request
from dataclasses import dataclass, field, asdict
from datetime import date, datetime
from pathlib import Path
from typing import Optional

logger = logging.getLogger(__name__)


# ── Data Classes ───────────────────────────────────────────────────────

@dataclass
class HoldingRecord:
    """A single holding within an ETF on a specific date."""
    ticker: str              # holding's stock ticker (e.g., "AAPL")
    name: str                # company name
    weight: float            # portfolio weight 0.0–1.0
    shares: int = 0          # number of shares held
    market_value: float = 0  # market value in USD
    sector: str = ""         # GICS sector from provider
    naics_codes: list[str] = field(default_factory=list)  # mapped NAICS
    asset_class: str = ""    # Equity, Bond, Cash, etc.
    sedol: str = ""          # SEDOL identifier
    isin: str = ""           # ISIN identifier
    exchange: str = ""       # exchange code

    def to_dict(self) -> dict:
        return asdict(self)

    @classmethod
    def from_dict(cls, d: dict) -> "HoldingRecord":
        return cls(**{k: v for k, v in d.items() if k in cls.__dataclass_fields__})


@dataclass
class HoldingsSnapshot:
    """Complete holdings for one ETF on one date."""
    etf_ticker: str
    snapshot_date: date
    source: str              # "ishares_csv", "sec_nport", "manual"
    holdings: list[HoldingRecord]
    total_holdings: int = 0
    total_net_assets: float = 0
    fetch_timestamp: str = ""

    def __post_init__(self):
        if not self.total_holdings:
            self.total_holdings = len(self.holdings)
        if not self.fetch_timestamp:
            self.fetch_timestamp = datetime.utcnow().isoformat()

    @property
    def equity_holdings(self) -> list[HoldingRecord]:
        """Only equity holdings (exclude cash, futures, etc.)."""
        return [h for h in self.holdings
                if h.asset_class.upper() in ("EQUITY", "STOCK", "")
                and h.weight > 0]

    @property
    def top_holdings(self) -> list[HoldingRecord]:
        """Top 10 holdings by weight."""
        return sorted(self.holdings, key=lambda h: h.weight, reverse=True)[:10]

    def naics_exposure(self) -> dict[str, float]:
        """Calculate NAICS exposure from actual holdings.

        Returns {naics_code: total_weight} aggregated across all holdings.
        This replaces the static approximations in etf_exposure.py.
        """
        exposure: dict[str, float] = {}
        for h in self.equity_holdings:
            for naics in h.naics_codes:
                exposure[naics] = exposure.get(naics, 0.0) + h.weight
        return exposure

    def to_dict(self) -> dict:
        return {
            "etf_ticker": self.etf_ticker,
            "snapshot_date": self.snapshot_date.isoformat(),
            "source": self.source,
            "total_holdings": self.total_holdings,
            "total_net_assets": self.total_net_assets,
            "fetch_timestamp": self.fetch_timestamp,
            "holdings": [h.to_dict() for h in self.holdings],
        }

    @classmethod
    def from_dict(cls, d: dict) -> "HoldingsSnapshot":
        holdings = [HoldingRecord.from_dict(h) for h in d.get("holdings", [])]
        return cls(
            etf_ticker=d["etf_ticker"],
            snapshot_date=date.fromisoformat(d["snapshot_date"]),
            source=d.get("source", "unknown"),
            holdings=holdings,
            total_holdings=d.get("total_holdings", len(holdings)),
            total_net_assets=d.get("total_net_assets", 0),
            fetch_timestamp=d.get("fetch_timestamp", ""),
        )


# ── iShares CSV Parser ─────────────────────────────────────────────────
#
# BlackRock publishes daily holdings CSVs for all iShares ETFs.
# URL pattern: https://www.ishares.com/us/products/{fund_id}/ishares-...
# The CSV has a header section (fund info) then holdings data.

# Fund IDs for iShares ETFs in our universe
ISHARES_FUND_IDS: dict[str, str] = {
    "IBB":  "239714",   # iShares Biotechnology
    "IHI":  "239516",   # iShares Medical Devices
    "IHF":  "239519",   # iShares Healthcare Providers
    "ITA":  "239502",   # iShares Aerospace & Defense
    "IYT":  "239501",   # iShares Transportation
    "IYZ":  "239525",   # iShares Telecommunications
    "IAI":  "239508",   # iShares Broker-Dealers
    "ICLN": "239738",   # iShares Clean Energy
}

ISHARES_CSV_URL = (
    "https://www.ishares.com/us/products/{fund_id}/"
    "1467271812596.ajax?fileType=csv&fileName={ticker}-holdings&dataType=fund"
)


def _fetch_url(url: str, timeout: int = 30) -> str:
    """Fetch URL content as string."""
    req = urllib.request.Request(url, headers={
        "User-Agent": "Legalize/0.1 (holdings-collector)",
        "Accept": "text/csv,text/plain,*/*",
    })
    with urllib.request.urlopen(req, timeout=timeout) as resp:
        return resp.read().decode("utf-8", errors="replace")


def _fetch_url_bytes(url: str, timeout: int = 30) -> bytes:
    """Fetch URL content as raw bytes (for XLSX files)."""
    req = urllib.request.Request(url, headers={
        "User-Agent": "Legalize/0.1 (holdings-collector)",
        "Accept": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet,*/*",
    })
    with urllib.request.urlopen(req, timeout=timeout) as resp:
        return resp.read()


def parse_ishares_csv(csv_text: str, etf_ticker: str) -> HoldingsSnapshot:
    """Parse an iShares holdings CSV into a HoldingsSnapshot.

    iShares CSVs have a metadata header section followed by holdings data.
    The holdings section starts after a row containing column headers like
    "Ticker,Name,Sector,Asset Class,Market Value,Weight (%),..."
    """
    lines = csv_text.strip().split("\n")

    # Find the header row (contains "Ticker" and "Weight")
    header_idx = -1
    for i, line in enumerate(lines):
        if "Ticker" in line and ("Weight" in line or "Market Value" in line):
            header_idx = i
            break

    if header_idx == -1:
        # Try alternate: look for "Name" and "Shares"
        for i, line in enumerate(lines):
            if "Name" in line and "Shares" in line:
                header_idx = i
                break

    if header_idx == -1:
        logger.warning(f"Could not find header row in iShares CSV for {etf_ticker}")
        return HoldingsSnapshot(
            etf_ticker=etf_ticker,
            snapshot_date=date.today(),
            source="ishares_csv",
            holdings=[],
        )

    # Parse as CSV from header row onward
    data_text = "\n".join(lines[header_idx:])
    reader = csv.DictReader(io.StringIO(data_text))

    holdings = []
    for row in reader:
        # Skip empty rows or summary rows
        ticker = row.get("Ticker", "").strip()
        name = row.get("Name", "").strip()
        if not name or name == "-":
            continue

        # Parse weight (could be "1.23" or "1.23%")
        weight_str = row.get("Weight (%)", row.get("Weight", "0"))
        weight_str = weight_str.replace("%", "").replace(",", "").strip()
        try:
            weight = float(weight_str) / 100.0  # convert percentage to decimal
        except (ValueError, TypeError):
            weight = 0.0

        # Parse shares
        shares_str = row.get("Shares", "0").replace(",", "").strip()
        try:
            shares = int(float(shares_str))
        except (ValueError, TypeError):
            shares = 0

        # Parse market value
        mv_str = row.get("Market Value", "0").replace(",", "").replace("$", "").strip()
        try:
            market_value = float(mv_str)
        except (ValueError, TypeError):
            market_value = 0.0

        sector = row.get("Sector", "").strip()
        asset_class = row.get("Asset Class", "").strip()
        sedol = row.get("SEDOL", "").strip()
        isin = row.get("ISIN", "").strip()
        exchange = row.get("Exchange", "").strip()

        holdings.append(HoldingRecord(
            ticker=ticker if ticker != "-" else "",
            name=name,
            weight=weight,
            shares=shares,
            market_value=market_value,
            sector=sector,
            asset_class=asset_class,
            sedol=sedol,
            isin=isin,
            exchange=exchange,
        ))

    # Extract fund-level metadata from header section
    total_assets = 0.0
    for line in lines[:header_idx]:
        if "Net Assets" in line or "Fund Holdings" in line:
            numbers = re.findall(r"[\d,]+\.?\d*", line)
            if numbers:
                try:
                    total_assets = float(numbers[0].replace(",", ""))
                except ValueError:
                    pass

    return HoldingsSnapshot(
        etf_ticker=etf_ticker,
        snapshot_date=date.today(),
        source="ishares_csv",
        holdings=holdings,
        total_net_assets=total_assets,
    )


def parse_spdr_xlsx(xlsx_data: bytes, etf_ticker: str) -> HoldingsSnapshot:
    """Parse a State Street SPDR ETF holdings XLSX file.

    SPDR XLSX format:
      - First few rows: metadata (fund name, as-of date, etc.)
      - Header row: Name, Ticker, Identifier, SEDOL, Weight, Sector,
                    Shares Held, Local Currency, etc.
      - Data rows: one per holding

    Column names vary slightly between funds, so we use fuzzy matching.
    """
    try:
        from openpyxl import load_workbook
    except ImportError:
        logger.error("openpyxl required for SPDR XLSX parsing: pip install openpyxl")
        return None

    wb = load_workbook(io.BytesIO(xlsx_data), read_only=True, data_only=True)
    ws = wb.active

    # Read all rows into a list for easier processing
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    if not rows:
        return None

    # Find the header row (contains "Name" and "Weight" or "Ticker")
    header_idx = None
    col_map: dict[str, int] = {}

    # Column name patterns (fuzzy matching)
    COL_PATTERNS = {
        "name": ["name", "holding name", "security name", "description"],
        "ticker": ["ticker", "symbol", "stock ticker"],
        "weight": ["weight", "weight (%)", "% of net assets", "portfolio %"],
        "sector": ["sector", "gics sector", "industry sector"],
        "shares": ["shares", "shares held", "quantity"],
        "market_value": ["market value", "notional value", "value"],
        "sedol": ["sedol", "sedol id"],
        "isin": ["isin", "isin id"],
        "identifier": ["identifier", "cusip", "id"],
    }

    for i, row in enumerate(rows):
        if row is None:
            continue
        # Convert to lowercase strings for matching
        cells = [str(c).strip().lower() if c is not None else "" for c in row]

        # Check if this looks like a header row
        has_name = any(any(pat in cell for pat in COL_PATTERNS["name"]) for cell in cells)
        has_weight = any(any(pat in cell for pat in COL_PATTERNS["weight"]) for cell in cells)

        if has_name and has_weight:
            header_idx = i
            # Build column mapping
            for j, cell in enumerate(cells):
                for field_name, patterns in COL_PATTERNS.items():
                    if any(pat in cell for pat in patterns):
                        if field_name not in col_map:
                            col_map[field_name] = j
            break

    if header_idx is None:
        logger.warning(f"Could not find header row in SPDR XLSX for {etf_ticker}")
        return None

    # Parse data rows
    holdings: list[HoldingRecord] = []
    for row in rows[header_idx + 1:]:
        if row is None or all(c is None for c in row):
            continue

        def get_val(field_name, default=""):
            idx = col_map.get(field_name)
            if idx is not None and idx < len(row) and row[idx] is not None:
                return row[idx]
            return default

        name = str(get_val("name", "")).strip()
        if not name or name.lower() in ("", "total", "cash", "other", "--"):
            continue

        ticker_val = str(get_val("ticker", "")).strip()
        weight_raw = get_val("weight", 0)
        sector_val = str(get_val("sector", "")).strip()
        shares_raw = get_val("shares", 0)
        mv_raw = get_val("market_value", 0)
        sedol_val = str(get_val("sedol", "")).strip()
        isin_val = str(get_val("isin", "")).strip()

        # Parse weight (might be string with % or numeric)
        try:
            if isinstance(weight_raw, str):
                weight = float(weight_raw.replace("%", "").replace(",", "").strip())
            else:
                weight = float(weight_raw)
            # SPDR weights are in percentage (0-100), normalize to 0-1
            if weight > 1.0:
                weight = weight / 100.0
        except (ValueError, TypeError):
            weight = 0.0

        if weight <= 0:
            continue

        # Parse shares
        try:
            if isinstance(shares_raw, str):
                shares = int(float(shares_raw.replace(",", "").strip()))
            else:
                shares = int(float(shares_raw)) if shares_raw else 0
        except (ValueError, TypeError):
            shares = 0

        # Parse market value
        try:
            if isinstance(mv_raw, str):
                market_value = float(mv_raw.replace(",", "").replace("$", "").strip())
            else:
                market_value = float(mv_raw) if mv_raw else 0.0
        except (ValueError, TypeError):
            market_value = 0.0

        # Map sector to NAICS
        naics_codes = gics_sector_to_naics(sector_val) if sector_val else []

        holdings.append(HoldingRecord(
            ticker=ticker_val if ticker_val and ticker_val != "--" else "",
            name=name,
            weight=weight,
            shares=shares,
            market_value=market_value,
            sector=sector_val,
            naics_codes=naics_codes,
            asset_class="Equity",
            sedol=sedol_val if sedol_val and sedol_val != "--" else "",
            isin=isin_val if isin_val and isin_val != "--" else "",
        ))

    if not holdings:
        logger.warning(f"No holdings parsed from SPDR XLSX for {etf_ticker}")
        return None

    # Normalize weights if they don't sum to ~1.0
    weight_sum = sum(h.weight for h in holdings)
    if weight_sum > 0 and abs(weight_sum - 1.0) > 0.1:
        for h in holdings:
            h.weight = h.weight / weight_sum

    return HoldingsSnapshot(
        etf_ticker=etf_ticker,
        snapshot_date=date.today(),
        source="spdr_xlsx",
        holdings=holdings,
    )


# ── SPDR Holdings Parser ──────────────────────────────────────────────
#
# State Street publishes holdings for SPDR ETFs (XLE, XLF, XLV, etc.)
# URL: https://www.ssga.com/us/en/intermediary/etfs/library-content/
#      products/fund-data/etfs/us/holdings-daily-us-en-{ticker}.xlsx

SPDR_TICKERS = [
    "XLE", "XLF", "XLV", "XLK", "XLI", "XLU", "XLP", "XLY",
    "XLC", "XLRE", "XLB", "XBI", "XHB", "XME", "XOP", "XRT", "KRE", "KIE",
]

SPDR_HOLDINGS_URL = (
    "https://www.ssga.com/us/en/intermediary/etfs/library-content/"
    "products/fund-data/etfs/us/holdings-daily-us-en-{ticker}.xlsx"
)


# ── SIC → NAICS Crosswalk ─────────────────────────────────────────────
#
# Many ETF providers classify holdings by GICS sector, not NAICS.
# We need to map holdings to NAICS for our exposure engine.
# This is a simplified sector→NAICS mapping for v1.

_GICS_SECTOR_TO_NAICS: dict[str, list[str]] = {
    "Energy": ["211", "213", "324"],
    "Materials": ["212", "325", "331", "332", "322"],
    "Industrials": ["236", "332", "333", "481", "484", "561"],
    "Consumer Discretionary": ["441", "448", "454", "3361", "721"],
    "Consumer Staples": ["311", "312", "445", "446"],
    "Health Care": ["3254", "3391", "621", "622"],
    "Financials": ["522", "523", "524", "525"],
    "Information Technology": ["334", "5112", "5415", "518"],
    "Communication Services": ["517", "515", "518", "519"],
    "Utilities": ["2211", "2212", "2213"],
    "Real Estate": ["531", "525"],
}


def gics_sector_to_naics(sector: str) -> list[str]:
    """Map a GICS sector name to probable NAICS codes."""
    # Normalize
    normalized = sector.strip().title()
    # Direct match
    if normalized in _GICS_SECTOR_TO_NAICS:
        return _GICS_SECTOR_TO_NAICS[normalized]
    # Fuzzy: try contains
    for key, codes in _GICS_SECTOR_TO_NAICS.items():
        if key.lower() in normalized.lower() or normalized.lower() in key.lower():
            return codes
    return []


# ── Storage Engine ─────────────────────────────────────────────────────

class HoldingsStore:
    """Persistent storage for ETF holdings snapshots.

    Directory layout:
        {base_dir}/
            snapshots/
                {YYYY-MM-DD}/
                    {TICKER}.json
            naics_cache/
                {TICKER}.json    # company ticker → NAICS mapping
            metadata.json
    """

    def __init__(self, base_dir: str | Path = "holdings_data"):
        self.base_dir = Path(base_dir)
        self.snapshots_dir = self.base_dir / "snapshots"
        self.cache_dir = self.base_dir / "naics_cache"
        self._ensure_dirs()
        self._naics_cache: dict[str, list[str]] = {}
        self._load_naics_cache()

    def _ensure_dirs(self):
        self.snapshots_dir.mkdir(parents=True, exist_ok=True)
        self.cache_dir.mkdir(parents=True, exist_ok=True)

    def _load_naics_cache(self):
        """Load company ticker → NAICS code cache."""
        cache_file = self.cache_dir / "_all_naics.json"
        if cache_file.exists():
            try:
                self._naics_cache = json.loads(cache_file.read_text())
            except (json.JSONDecodeError, IOError):
                self._naics_cache = {}

    def _save_naics_cache(self):
        cache_file = self.cache_dir / "_all_naics.json"
        cache_file.write_text(json.dumps(self._naics_cache, indent=2))

    def get_naics_for_ticker(self, ticker: str) -> list[str]:
        """Get cached NAICS codes for a company ticker."""
        return self._naics_cache.get(ticker, [])

    def set_naics_for_ticker(self, ticker: str, naics_codes: list[str]):
        """Cache NAICS codes for a company ticker."""
        self._naics_cache[ticker] = naics_codes
        self._save_naics_cache()

    def save_snapshot(self, snapshot: HoldingsSnapshot) -> Path:
        """Save a holdings snapshot to disk."""
        day_dir = self.snapshots_dir / snapshot.snapshot_date.isoformat()
        day_dir.mkdir(parents=True, exist_ok=True)
        filepath = day_dir / f"{snapshot.etf_ticker}.json"
        filepath.write_text(json.dumps(snapshot.to_dict(), indent=2))
        logger.info(f"Saved {snapshot.etf_ticker} snapshot to {filepath}")
        return filepath

    def load_snapshot(
        self, etf_ticker: str, snapshot_date: date
    ) -> Optional[HoldingsSnapshot]:
        """Load a specific snapshot from disk."""
        filepath = self.snapshots_dir / snapshot_date.isoformat() / f"{etf_ticker}.json"
        if not filepath.exists():
            return None
        try:
            data = json.loads(filepath.read_text())
            return HoldingsSnapshot.from_dict(data)
        except (json.JSONDecodeError, IOError, KeyError) as e:
            logger.error(f"Failed to load {filepath}: {e}")
            return None

    def get_available_dates(self, etf_ticker: Optional[str] = None) -> list[date]:
        """List all dates with snapshots."""
        dates = []
        if not self.snapshots_dir.exists():
            return dates
        for day_dir in sorted(self.snapshots_dir.iterdir()):
            if not day_dir.is_dir():
                continue
            try:
                d = date.fromisoformat(day_dir.name)
            except ValueError:
                continue
            if etf_ticker:
                if (day_dir / f"{etf_ticker}.json").exists():
                    dates.append(d)
            else:
                dates.append(d)
        return dates

    def get_latest_snapshot(self, etf_ticker: str) -> Optional[HoldingsSnapshot]:
        """Get the most recent snapshot for an ETF."""
        dates = self.get_available_dates(etf_ticker)
        if not dates:
            return None
        return self.load_snapshot(etf_ticker, dates[-1])

    def get_snapshot_history(
        self,
        etf_ticker: str,
        start_date: Optional[date] = None,
        end_date: Optional[date] = None,
    ) -> list[HoldingsSnapshot]:
        """Get all snapshots for an ETF within a date range."""
        dates = self.get_available_dates(etf_ticker)
        if start_date:
            dates = [d for d in dates if d >= start_date]
        if end_date:
            dates = [d for d in dates if d <= end_date]
        snapshots = []
        for d in dates:
            snap = self.load_snapshot(etf_ticker, d)
            if snap:
                snapshots.append(snap)
        return snapshots

    def summary(self) -> dict:
        """Storage summary stats."""
        all_dates = self.get_available_dates()
        etf_counts: dict[str, int] = {}
        for d in all_dates:
            day_dir = self.snapshots_dir / d.isoformat()
            for f in day_dir.glob("*.json"):
                ticker = f.stem
                etf_counts[ticker] = etf_counts.get(ticker, 0) + 1
        return {
            "total_dates": len(all_dates),
            "date_range": (
                f"{all_dates[0].isoformat()} to {all_dates[-1].isoformat()}"
                if all_dates else "none"
            ),
            "etfs_tracked": len(etf_counts),
            "etf_snapshot_counts": etf_counts,
            "naics_cache_size": len(self._naics_cache),
        }


# ── Holdings Collector ─────────────────────────────────────────────────

class HoldingsCollector:
    """Orchestrates daily ETF holdings collection.

    Usage:
        collector = HoldingsCollector(store=HoldingsStore("./holdings_data"))

        # Collect all ETFs
        results = collector.collect_all()

        # Collect specific ETF
        snapshot = collector.collect_one("XLE")
    """

    # All ETFs we track
    ALL_ETFS = sorted(set(list(ISHARES_FUND_IDS.keys()) + SPDR_TICKERS + [
        "HACK", "CIBR", "JETS", "SMH", "OIH", "TAN", "LIT",
        "MOO", "PHO", "IYZ", "IYT",
    ]))

    def __init__(
        self,
        store: Optional[HoldingsStore] = None,
        throttle: float = 1.0,
    ):
        self.store = store or HoldingsStore()
        self.throttle = throttle

    def collect_one(
        self,
        etf_ticker: str,
        force: bool = False,
    ) -> Optional[HoldingsSnapshot]:
        """Collect holdings for a single ETF.

        Args:
            etf_ticker: ETF ticker symbol
            force: Re-collect even if today's snapshot exists
        """
        today = date.today()

        # Check if we already have today's snapshot
        if not force:
            existing = self.store.load_snapshot(etf_ticker, today)
            if existing:
                logger.info(f"{etf_ticker}: already have today's snapshot")
                return existing

        snapshot = None

        # Try iShares CSV first
        if etf_ticker in ISHARES_FUND_IDS:
            snapshot = self._fetch_ishares(etf_ticker)

        # Try SPDR
        if not snapshot and etf_ticker in SPDR_TICKERS:
            snapshot = self._fetch_spdr(etf_ticker)

        # Fallback: create a minimal snapshot from our static data
        if not snapshot:
            snapshot = self._static_fallback(etf_ticker)

        if snapshot:
            # Enrich holdings with NAICS codes from cache/sector mapping
            self._enrich_naics(snapshot)
            # Save
            self.store.save_snapshot(snapshot)

        return snapshot

    def collect_all(self, force: bool = False) -> list[HoldingsSnapshot]:
        """Collect holdings for all tracked ETFs."""
        results = []
        for ticker in self.ALL_ETFS:
            try:
                snap = self.collect_one(ticker, force=force)
                if snap:
                    results.append(snap)
                time.sleep(self.throttle)
            except Exception as e:
                logger.error(f"Failed to collect {ticker}: {e}")
        return results

    def _fetch_ishares(self, ticker: str) -> Optional[HoldingsSnapshot]:
        """Fetch holdings from iShares CSV."""
        fund_id = ISHARES_FUND_IDS.get(ticker)
        if not fund_id:
            return None
        url = ISHARES_CSV_URL.format(fund_id=fund_id, ticker=ticker)
        try:
            csv_text = _fetch_url(url)
            return parse_ishares_csv(csv_text, ticker)
        except Exception as e:
            logger.error(f"iShares fetch failed for {ticker}: {e}")
            return None

    def _fetch_spdr(self, ticker: str) -> Optional[HoldingsSnapshot]:
        """Fetch and parse SPDR ETF holdings from State Street XLSX.

        State Street publishes daily holdings as XLSX files:
        https://www.ssga.com/.../holdings-daily-us-en-{ticker}.xlsx

        XLSX format (typical):
          Row 1-4: Header metadata (fund name, date, etc.)
          Row 5:   Column headers (Name, Ticker, Identifier, SEDOL,
                   Weight, Sector, Shares Held, Local Currency, etc.)
          Row 6+:  Holdings data
        """
        url = SPDR_HOLDINGS_URL.format(ticker=ticker.lower())
        try:
            xlsx_bytes = _fetch_url_bytes(url)
            return parse_spdr_xlsx(xlsx_bytes, ticker)
        except Exception as e:
            logger.error(f"SPDR fetch failed for {ticker}: {e}")
            return None

    def _static_fallback(self, ticker: str) -> Optional[HoldingsSnapshot]:
        """Create a snapshot from our static ETF data.

        This is the fallback when we can't fetch live data.
        It creates a synthetic snapshot matching our static exposure weights.
        """
        from .etf_exposure import ETFExposureEngine
        engine = ETFExposureEngine()
        profile = engine.get_profile(ticker)
        if not profile:
            return None

        # Create synthetic holdings from NAICS exposure weights
        holdings = []
        for naics, weight in profile.naics_exposure.items():
            holdings.append(HoldingRecord(
                ticker=f"NAICS-{naics}",
                name=f"NAICS {naics} aggregate",
                weight=weight,
                naics_codes=[naics],
                asset_class="Equity",
            ))

        return HoldingsSnapshot(
            etf_ticker=ticker,
            snapshot_date=date.today(),
            source="static_fallback",
            holdings=holdings,
        )

    def _enrich_naics(self, snapshot: HoldingsSnapshot):
        """Add NAICS codes to holdings from cache and sector mapping."""
        for holding in snapshot.holdings:
            # Skip if already has NAICS
            if holding.naics_codes:
                continue

            # Check cache
            if holding.ticker:
                cached = self.store.get_naics_for_ticker(holding.ticker)
                if cached:
                    holding.naics_codes = cached
                    continue

            # Map from GICS sector
            if holding.sector:
                naics = gics_sector_to_naics(holding.sector)
                if naics:
                    holding.naics_codes = naics
                    # Cache it
                    if holding.ticker:
                        self.store.set_naics_for_ticker(holding.ticker, naics)
